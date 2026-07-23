# reports/views.py
from datetime import datetime, timedelta, date
from decimal import Decimal

from django.views.generic import ListView, View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count, F, Q
from django.utils import timezone

from inventory.models import Item, Stock, StockMovement, Warehouse, Category
from procurement.models import PurchaseOrder, GoodsReceipt
from operations.models import Project, MaterialRequest
from assets.models import Asset
from accounts.models import User

# ─── Imports for all modules ──────────────────────────────
from sales.models import Invoice, Payment, SalesOrder
from finance.models import Account, JournalLine, Expense
from hr.models import Employee, LeaveRequest, Attendance
from company_settings.services import get_setting


# ─────────────────────────────────────────
# DASHBOARD VIEW
# ─────────────────────────────────────────

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.now().date()

        # ─── Base Context (always present) ──────────────────
        ctx.update({
            'today': today,
            'user_role': user.role.name if user.role else None,
            'currency_symbol': get_setting('CURRENCY_SYMBOL', 'MK'),
        })

        # ─── Inventory Stats ────────────────────────────────
        if user.has_perm('inventory.view_stock_report') or user.has_perm('inventory.view_item'):
            ctx.update(self._get_inventory_stats(today))

        # ─── Sales Stats ────────────────────────────────────
        if user.has_perm('sales.view_invoice') or user.has_perm('sales.view_salesorder'):
            ctx.update(self._get_sales_stats(today))

        # ─── Finance Stats ──────────────────────────────────
        if user.has_perm('finance.view_account') or user.has_perm('finance.view_expense'):
            ctx.update(self._get_finance_stats(today))

        # ─── HR Stats ──────────────────────────────────────
        if user.has_perm('hr.view_employee') or user.has_perm('hr.view_attendance'):
            ctx.update(self._get_hr_stats(today))

        # ─── Procurement Stats ──────────────────────────────
        if user.has_perm('procurement.view_purchaseorder') or user.has_perm('procurement.view_purchaserequest'):
            ctx.update(self._get_procurement_stats(today))

        # ─── Operations Stats ────────────────────────────────
        if user.has_perm('operations.view_project') or user.has_perm('operations.view_materialrequest'):
            ctx.update(self._get_operations_stats(today))

        # ─── Assets Stats ────────────────────────────────────
        if user.has_perm('assets.view_asset'):
            ctx.update(self._get_assets_stats(today))

        # ─── Recent Movements ────────────────────────────────
        if user.has_perm('inventory.view_stockmovement'):
            ctx['recent_movements'] = StockMovement.objects.select_related(
                'item', 'warehouse', 'created_by'
            ).order_by('-created_at')[:8]

        # ─── Chart Data ──────────────────────────────────────
        ctx.update(self._get_chart_data(today))

        # ─── Role‑specific Charts ────────────────────────────
        if user.has_perm('finance.view_account') or user.has_perm('finance.view_expense'):
            ctx.update(self._get_finance_chart_data(today))

        if user.has_perm('hr.view_employee') or user.has_perm('hr.view_attendance'):
            ctx.update(self._get_hr_chart_data(today))

        if user.has_perm('procurement.view_purchaseorder') or user.has_perm('procurement.view_goodsreceipt'):
            ctx.update(self._get_procurement_chart_data(today))

        return ctx

    # ─── Helper Methods ──────────────────────────────────────

    def _get_inventory_stats(self, today):
        return {
            'total_items': Item.objects.filter(is_active=True).count(),
            'total_warehouses': Warehouse.objects.filter(is_active=True).count(),
            'low_stock_items': Stock.objects.filter(
                quantity__lte=F('item__minimum_stock'),
                item__minimum_stock__gt=0
            ).select_related('item', 'warehouse')[:5],
            'low_stock_count': Stock.objects.filter(
                quantity__lte=F('item__minimum_stock'),
                item__minimum_stock__gt=0
            ).count(),
            'stock_value': Stock.objects.aggregate(
                total=Sum(F('quantity') * F('item__unit_cost'))
            )['total'] or 0,
            'received_today': StockMovement.objects.filter(
                movement_type='IN', created_at__date=today
            ).count(),
            'issued_today': StockMovement.objects.filter(
                movement_type='OUT', created_at__date=today
            ).count(),
        }

    def _get_sales_stats(self, today):
        month_start = today.replace(day=1)
        return {
            'today_sales': Payment.objects.filter(
                payment_date__date=today
            ).aggregate(total=Sum('amount'))['total'] or 0,
            'monthly_sales': Payment.objects.filter(
                payment_date__date__gte=month_start,
                payment_date__date__lte=today
            ).aggregate(total=Sum('amount'))['total'] or 0,
            'outstanding_invoices': Invoice.objects.filter(
                status__in=('sent', 'partially_paid')
            ).count(),
            'pending_sales_orders': SalesOrder.objects.filter(
                status__in=('draft', 'pending_approval')
            ).count(),
            'ready_to_invoice_count': SalesOrder.objects.filter(
                status='approved',
                invoices__isnull=True
            ).count(),
        }

    def _get_finance_stats(self, today):
        month_start = today.replace(day=1)
        cash_code = get_setting('PAYROLL_CASH_ACCOUNT', '1000')
        return {
            'total_revenue': Payment.objects.filter(
                payment_date__date__gte=month_start,
                payment_date__date__lte=today
            ).aggregate(total=Sum('amount'))['total'] or 0,
            'total_expenses': Expense.objects.filter(
                status='paid',
                created_at__date__gte=month_start,
                created_at__date__lte=today
            ).aggregate(total=Sum('amount'))['total'] or 0,
            'net_profit': Payment.objects.filter(
                payment_date__date__gte=month_start,
                payment_date__date__lte=today
            ).aggregate(total=Sum('amount'))['total'] or 0 -
            Expense.objects.filter(
                status='paid',
                created_at__date__gte=month_start,
                created_at__date__lte=today
            ).aggregate(total=Sum('amount'))['total'] or 0,
            'cash_balance': JournalLine.objects.filter(
                account__code=cash_code
            ).aggregate(balance=Sum('debit') - Sum('credit'))['balance'] or 0,
            'pending_expenses': Expense.objects.filter(
                status__in=('pending_approval', 'approved')
            ).count(),
        }

    def _get_hr_stats(self, today):
        return {
            'total_employees': Employee.objects.filter(is_active=True).count(),
            'pending_leave_requests': LeaveRequest.objects.filter(status='pending').count(),
            'employees_on_leave_today': LeaveRequest.objects.filter(
                status='approved',
                start_date__lte=today,
                end_date__gte=today
            ).count(),
            'clocked_in_today': Attendance.objects.filter(date=today, clock_in__isnull=False).count(),
        }

    def _get_procurement_stats(self, today):
        return {
            'pending_orders': PurchaseOrder.objects.filter(
                status__in=('draft', 'sent')
            ).count(),
            'pending_pr_approvals': MaterialRequest.objects.filter(
                status='submitted'
            ).count(),
        }

    def _get_operations_stats(self, today):
        return {
            'active_projects': Project.objects.filter(status='active').count(),
        }

    def _get_assets_stats(self, today):
        return {
            'assets_assigned': Asset.objects.filter(status='assigned').count(),
            'assets_due_maintenance': Asset.objects.filter(
                next_maintenance_date__lte=today + timedelta(days=7),
                next_maintenance_date__isnull=False,
                status__in=('available', 'assigned')
            ).count(),
        }

    def _get_chart_data(self, today):
        # Stock movement chart
        chart_in = []
        chart_out = []
        sales_chart = []
        labels = []

        for i in range(6, -1, -1):
            day = today - timedelta(days=i)
            labels.append(day.strftime('%d %b'))
            chart_in.append(StockMovement.objects.filter(
                movement_type='IN', created_at__date=day
            ).count())
            chart_out.append(StockMovement.objects.filter(
                movement_type='OUT', created_at__date=day
            ).count())
            sales_chart.append(float(Payment.objects.filter(
                payment_date__date=day
            ).aggregate(total=Sum('amount'))['total'] or 0))

        return {
            'chart_labels': labels,
            'movement_chart_in': chart_in,
            'movement_chart_out': chart_out,
            'sales_chart': sales_chart,
        }

    def _get_finance_chart_data(self, today):
        month_start = today.replace(day=1)
        # Expense breakdown (top 5 categories)
        expenses_by_category = Expense.objects.filter(
            status='paid',
            created_at__date__gte=month_start,
            created_at__date__lte=today
        ).values('category').annotate(total=Sum('amount')).order_by('-total')[:5]
        expense_labels = [e['category'].title() for e in expenses_by_category]
        expense_values = [float(e['total']) for e in expenses_by_category]

        # Revenue trend (last 7 days)
        revenue_trend = []
        for i in range(6, -1, -1):
            day = today - timedelta(days=i)
            revenue_trend.append(float(Payment.objects.filter(
                payment_date__date=day
            ).aggregate(total=Sum('amount'))['total'] or 0))

        return {
            'expense_labels': expense_labels,
            'expense_values': expense_values,
            'revenue_trend': revenue_trend,
        }

    def _get_hr_chart_data(self, today):
        # Leave distribution (by leave type)
        leave_by_type = LeaveRequest.objects.filter(
            status='approved',
            start_date__lte=today,
            end_date__gte=today
        ).values('leave_type__name').annotate(count=Count('id'))
        leave_labels = [l['leave_type__name'] for l in leave_by_type]
        leave_values = [l['count'] for l in leave_by_type]

        # Attendance status today
        total_emp = Employee.objects.filter(is_active=True).count()
        clocked_in = Attendance.objects.filter(date=today, clock_in__isnull=False).count()
        on_leave = LeaveRequest.objects.filter(
            status='approved',
            start_date__lte=today,
            end_date__gte=today
        ).count()
        absent = total_emp - clocked_in - on_leave

        return {
            'leave_labels': leave_labels,
            'leave_values': leave_values,
            'attendance_clocked': clocked_in,
            'attendance_on_leave': on_leave,
            'attendance_absent': absent,
        }

    def _get_procurement_chart_data(self, today):
        # PO status breakdown
        po_status = PurchaseOrder.objects.values('status').annotate(count=Count('id'))
        po_labels = [s['status'].title() for s in po_status]
        po_values = [s['count'] for s in po_status]

        # Received goods over last 7 days
        grn_trend = []
        for i in range(6, -1, -1):
            day = today - timedelta(days=i)
            grn_trend.append(GoodsReceipt.objects.filter(
                received_at__date=day
            ).count())

        return {
            'po_labels': po_labels,
            'po_values': po_values,
            'grn_trend': grn_trend,
        }


class DashboardStatsPartialView(LoginRequiredMixin, View):
    """HTMX endpoint — refreshes stat cards every 30s."""
    def get(self, request):
        # Reuse the dashboard logic but return only the stats partial
        view = DashboardView()
        context = view.get_context_data()
        return render(request, 'reports/partials/dashboard_stats.html', context)


# ─────────────────────────────────────────
# INVENTORY REPORTS
# ─────────────────────────────────────────

class InventoryReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/inventory_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        warehouse_id = self.request.GET.get('warehouse')
        category_id = self.request.GET.get('category')
        show = self.request.GET.get('show', 'all')

        stocks = Stock.objects.select_related(
            'item', 'item__category', 'item__unit', 'warehouse'
        ).order_by('item__name')

        if warehouse_id:
            stocks = stocks.filter(warehouse_id=warehouse_id)
        if category_id:
            stocks = stocks.filter(item__category_id=category_id)
        if show == 'low':
            stocks = stocks.filter(
                quantity__lte=F('item__minimum_stock'),
                item__minimum_stock__gt=0
            )
        elif show == 'zero':
            stocks = stocks.filter(quantity=0)

        total_value = stocks.aggregate(
            total=Sum(F('quantity') * F('item__unit_cost'))
        )['total'] or 0

        ctx.update({
            'stocks': stocks,
            'total_value': total_value,
            'warehouses': Warehouse.objects.filter(is_active=True),
            'categories': Category.objects.all(),
            'selected_warehouse': warehouse_id or '',
            'selected_category': category_id or '',
            'show': show,
        })
        return ctx


class StockMovementReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/movement_report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.now().date()

        date_from = self.request.GET.get('date_from', str(today - timedelta(days=30)))
        date_to = self.request.GET.get('date_to', str(today))
        movement_type = self.request.GET.get('type', '')
        warehouse_id = self.request.GET.get('warehouse', '')

        movements = StockMovement.objects.select_related(
            'item', 'item__unit', 'warehouse', 'created_by'
        ).filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).order_by('-created_at')

        if movement_type:
            movements = movements.filter(movement_type=movement_type)
        if warehouse_id:
            movements = movements.filter(warehouse_id=warehouse_id)

        ctx.update({
            'movements': movements,
            'date_from': date_from,
            'date_to': date_to,
            'movement_types': StockMovement.MOVEMENT_TYPES,
            'warehouses': Warehouse.objects.filter(is_active=True),
            'selected_type': movement_type,
            'selected_warehouse': warehouse_id,
            'total_in': movements.filter(
                movement_type='IN'
            ).aggregate(t=Sum('quantity'))['t'] or 0,
            'total_out': movements.filter(
                movement_type='OUT'
            ).aggregate(t=Sum('quantity'))['t'] or 0,
        })
        return ctx


class ProjectConsumptionReportView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/project_consumption.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        project_id = self.request.GET.get('project')

        projects = Project.objects.filter(status='active')
        selected_project = None
        consumption = []

        if project_id:
            selected_project = Project.objects.filter(pk=project_id).first()
            if selected_project:
                for mr in selected_project.material_requests.filter(
                    status__in=('issued', 'partially_issued')
                ).prefetch_related('items__item__unit'):
                    for item in mr.items.all():
                        if item.quantity_issued > 0:
                            consumption.append({
                                'item': item.item,
                                'quantity': item.quantity_issued,
                                'cost': item.quantity_issued * item.item.unit_cost,
                                'reference': mr.reference,
                                'date': mr.created_at,
                            })

        ctx.update({
            'projects': projects,
            'selected_project': selected_project,
            'consumption': consumption,
            'total_cost': sum(c['cost'] for c in consumption),
        })
        return ctx


# ─────────────────────────────────────────
# FINANCE REPORTS
# ─────────────────────────────────────────

class IncomeStatementView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/income_statement.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.now().date()
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = today.replace(day=1)
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = today

        income = Payment.objects.filter(
            payment_date__date__gte=date_from,
            payment_date__date__lte=date_to
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        expenses = Expense.objects.filter(
            status='paid',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).values('category').annotate(total=Sum('amount')).order_by('category')

        total_expenses = sum(e['total'] for e in expenses) if expenses else Decimal('0.00')
        net_income = income - total_expenses

        ctx.update({
            'income': income,
            'expenses': expenses,
            'total_expenses': total_expenses,
            'net_income': net_income,
            'date_from': date_from,
            'date_to': date_to,
        })
        return ctx


class CashFlowView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/cash_flow.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.now().date()
        date_from = self.request.GET.get('date_from', str(today.replace(day=1)))
        date_to = self.request.GET.get('date_to', str(today))

        cash_in = Payment.objects.filter(
            payment_date__date__gte=date_from,
            payment_date__date__lte=date_to,
            payment_method='cash'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        cash_out = Expense.objects.filter(
            status='paid',
            payment_method='cash',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        net_cash_flow = cash_in - cash_out

        ctx.update({
            'cash_in': cash_in,
            'cash_out': cash_out,
            'net_cash_flow': net_cash_flow,
            'date_from': date_from,
            'date_to': date_to,
        })
        return ctx


class BalanceSheetView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/balance_sheet.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        cash_code = get_setting('PAYROLL_CASH_ACCOUNT', '1000')
        bank_code = '1010'
        paye_code = '2005'
        pension_code = '2006'

        def get_balance(account_code):
            balance = JournalLine.objects.filter(
                account__code=account_code
            ).aggregate(b=Sum('debit') - Sum('credit'))['b'] or Decimal('0.00')
            return balance

        cash = get_balance(cash_code)
        bank = get_balance(bank_code)
        paye_liability = get_balance(paye_code)
        pension_liability = get_balance(pension_code)

        total_assets = cash + bank
        total_liabilities = paye_liability + pension_liability
        equity = total_assets - total_liabilities

        ctx.update({
            'cash': cash,
            'bank': bank,
            'total_assets': total_assets,
            'paye_liability': paye_liability,
            'pension_liability': pension_liability,
            'total_liabilities': total_liabilities,
            'equity': equity,
        })
        return ctx


# ─────────────────────────────────────────
# EXCEL EXPORTS
# ─────────────────────────────────────────

class ExportInventoryExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1E293B",
            end_color="1E293B",
            fill_type="solid"
        )

        headers = [
            'SKU', 'Item Name', 'Category', 'Type',
            'Unit', 'Warehouse', 'Quantity',
            'Min Stock', 'Unit Cost (MWK)', 'Total Value (MWK)', 'Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        stocks = Stock.objects.select_related(
            'item', 'item__category', 'item__unit', 'warehouse'
        ).order_by('item__name')

        for row, stock in enumerate(stocks, 2):
            total_value = stock.quantity * stock.item.unit_cost
            status = "Low Stock" if stock.is_low_stock else "OK"
            ws.append([
                stock.item.sku,
                stock.item.name,
                stock.item.category.name,
                stock.item.get_item_type_display(),
                stock.item.unit.symbol,
                stock.warehouse.name,
                float(stock.quantity),
                float(stock.item.minimum_stock),
                float(stock.item.unit_cost),
                float(total_value),
                status,
            ])

        column_widths = [12, 35, 20, 18, 8, 20, 12, 12, 18, 18, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="inventory_report_{date.today()}.xlsx"'
        )
        wb.save(response)
        return response


class ExportMovementsExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Movements"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1E293B",
            end_color="1E293B",
            fill_type="solid"
        )

        headers = [
            'Date', 'Item', 'SKU', 'Type',
            'Quantity', 'Unit', 'Warehouse',
            'Reference', 'Recorded By'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        today = timezone.now().date()
        date_from = request.GET.get(
            'date_from', str(today - timedelta(days=30))
        )
        date_to = request.GET.get('date_to', str(today))

        movements = StockMovement.objects.select_related(
            'item', 'item__unit', 'warehouse', 'created_by'
        ).filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).order_by('-created_at')

        for movement in movements:
            ws.append([
                movement.created_at.strftime('%d/%m/%Y %H:%M'),
                movement.item.name,
                movement.item.sku,
                movement.movement_type,
                float(movement.quantity),
                movement.item.unit.symbol,
                movement.warehouse.name,
                movement.reference or '',
                movement.created_by.get_full_name(),
            ])

        column_widths = [18, 35, 12, 12, 12, 8, 20, 18, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="movements_{date_from}_to_{date_to}.xlsx"'
        )
        wb.save(response)
        return response


# ─────────────────────────────────────────
# PDF EXPORTS
# ─────────────────────────────────────────

class ExportInventoryPDFView(LoginRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="inventory_{date.today()}.pdf"'
        )

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=1.5*cm
        )

        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=6,
        )
        elements.append(
            Paragraph("Warehouse Management System", title_style)
        )
        elements.append(
            Paragraph(
                f"Inventory Report — {date.today().strftime('%d %B %Y')}",
                styles['Normal']
            )
        )
        elements.append(Spacer(1, 0.5*cm))

        data = [[
            'SKU', 'Item Name', 'Category',
            'Warehouse', 'Qty', 'Unit',
            'Min Stock', 'Unit Cost', 'Total Value', 'Status'
        ]]

        stocks = Stock.objects.select_related(
            'item', 'item__category',
            'item__unit', 'warehouse'
        ).order_by('item__name')

        total_value = 0
        for stock in stocks:
            value = stock.quantity * stock.item.unit_cost
            total_value += value
            data.append([
                stock.item.sku,
                stock.item.name[:35],
                stock.item.category.name[:20],
                stock.warehouse.name[:20],
                f"{stock.quantity:,.2f}",
                stock.item.unit.symbol,
                f"{stock.item.minimum_stock:,.2f}",
                f"MWK {stock.item.unit_cost:,.2f}",
                f"MWK {value:,.2f}",
                "⚠ Low" if stock.is_low_stock else "OK",
            ])

        data.append([
            '', '', '', 'TOTAL', '', '', '', '',
            f"MWK {total_value:,.2f}", ''
        ])

        table = Table(
            data,
            colWidths=[
                2.2*cm, 6*cm, 3.5*cm, 3.5*cm,
                1.8*cm, 1.4*cm, 2*cm, 3.2*cm, 3.5*cm, 1.8*cm
            ]
        )

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [colors.white, colors.HexColor('#F8FAFC')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))

        elements.append(table)
        doc.build(elements)
        return response


class GlobalSearchView(LoginRequiredMixin, View):
    def get(self, request):
        q = request.GET.get('q', '').strip()
        results = {}

        if len(q) >= 2:
            from inventory.models import Item, Warehouse
            from procurement.models import Supplier, PurchaseOrder
            from operations.models import Project

            results['items'] = Item.objects.filter(
                Q(name__icontains=q) | Q(sku__icontains=q)
            )[:5]

            results['warehouses'] = Warehouse.objects.filter(
                name__icontains=q
            )[:3]

            results['suppliers'] = Supplier.objects.filter(
                name__icontains=q
            )[:3]

            results['projects'] = Project.objects.filter(
                Q(name__icontains=q) | Q(code__icontains=q)
            )[:3]

            results['orders'] = PurchaseOrder.objects.filter(
                reference__icontains=q
            )[:3]

        from django.template.loader import render_to_string
        from django.http import HttpResponse
        html = render_to_string(
            'reports/partials/search_results.html',
            {'results': results, 'q': q},
            request=request
        )
        return HttpResponse(html)


# ─────────────────────────────────────────
# FINANCE REPORT EXPORTS (EXCEL & PDF)
# ─────────────────────────────────────────

class ExportIncomeStatementExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from company_settings.models import Company

        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        today = timezone.now().date()
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = today.replace(day=1)
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = today

        income = Payment.objects.filter(
            payment_date__date__gte=date_from,
            payment_date__date__lte=date_to
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        expenses = Expense.objects.filter(
            status='paid',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).values('category').annotate(total=Sum('amount')).order_by('category')
        total_expenses = sum(e['total'] for e in expenses) if expenses else Decimal('0.00')
        net_income = income - total_expenses

        company = Company.objects.first()
        company_name = company.name if company else "J&N WMS"
        currency = company.currency_symbol if company else "MWK"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        ws.merge_cells('A1:B1')
        ws['A1'] = company_name
        ws['A1'].font = title_font
        ws.merge_cells('A2:B2')
        ws['A2'] = "Income Statement"
        ws.merge_cells('A3:B3')
        ws['A3'] = f"{date_from.strftime('%d %b %Y')} - {date_to.strftime('%d %b %Y')}"
        ws['A3'].font = Font(italic=True)

        headers = ['Category', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        row = 6
        ws.cell(row=row, column=1, value="Revenue").font = bold_font
        ws.cell(row=row, column=2, value=income).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1
        ws.cell(row=row, column=1, value="Expenses").font = bold_font
        row += 1
        for exp in expenses:
            ws.cell(row=row, column=1, value=exp['category'].title())
            ws.cell(row=row, column=2, value=exp['total']).number_format = '#,##0.00'
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
            row += 1
        ws.cell(row=row, column=1, value="Total Expenses").font = bold_font
        ws.cell(row=row, column=2, value=total_expenses).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1
        ws.cell(row=row, column=1, value="Net Profit").font = bold_font
        net_cell = ws.cell(row=row, column=2, value=net_income)
        net_cell.number_format = '#,##0.00'
        net_cell.alignment = Alignment(horizontal='right')
        if net_income >= 0:
            net_cell.font = Font(bold=True, color="22C55E")
        else:
            net_cell.font = Font(bold=True, color="EF4444")

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="income_statement_{date_from}_{date_to}.xlsx"'
        )
        wb.save(response)
        return response


class ExportIncomeStatementPDFView(LoginRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from company_settings.models import Company

        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        today = timezone.now().date()
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        else:
            date_from = today.replace(day=1)
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        else:
            date_to = today

        income = Payment.objects.filter(
            payment_date__date__gte=date_from,
            payment_date__date__lte=date_to
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        expenses = Expense.objects.filter(
            status='paid',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).values('category').annotate(total=Sum('amount')).order_by('category')
        total_expenses = sum(e['total'] for e in expenses) if expenses else Decimal('0.00')
        net_income = income - total_expenses

        company = Company.objects.first()
        company_name = company.name if company else "J&N WMS"
        currency = company.currency_symbol if company else "MWK"

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="income_statement_{date_from}_{date_to}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=16, textColor=colors.HexColor('#1E293B'),
                                     spaceAfter=6)
        elements.append(Paragraph(company_name, title_style))
        elements.append(Paragraph("Income Statement", styles['Heading2']))
        elements.append(Paragraph(f"{date_from.strftime('%d %b %Y')} - {date_to.strftime('%d %b %Y')}",
                                  styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        data = [['Category', f'Amount ({currency})']]
        data.append(['Revenue', f"{income:,.2f}"])
        data.append(['', ''])
        data.append(['Expenses', ''])
        for exp in expenses:
            data.append([exp['category'].title(), f"{exp['total']:,.2f}"])
        data.append(['Total Expenses', f"{total_expenses:,.2f}"])
        data.append(['', ''])
        data.append(['Net Profit', f"{net_income:,.2f}"])

        table = Table(data, colWidths=[10*cm, 6*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

        doc.build(elements)
        return response


class ExportCashFlowExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from company_settings.models import Company

        today = timezone.now().date()
        date_from = request.GET.get('date_from', str(today.replace(day=1)))
        date_to = request.GET.get('date_to', str(today))

        cash_in = Payment.objects.filter(
            payment_date__date__gte=date_from,
            payment_date__date__lte=date_to,
            payment_method='cash'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        cash_out = Expense.objects.filter(
            status='paid',
            payment_method='cash',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        net_cash_flow = cash_in - cash_out

        company = Company.objects.first()
        currency = company.currency_symbol if company else "MWK"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:B1')
        ws['A1'] = company.name if company else "J&N WMS"
        ws['A1'].font = title_font
        ws.merge_cells('A2:B2')
        ws['A2'] = "Cash Flow"
        ws.merge_cells('A3:B3')
        ws['A3'] = f"{date_from} - {date_to}"
        ws['A3'].font = Font(italic=True)

        headers = ['Category', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        row = 6
        ws.cell(row=row, column=1, value="Cash In (Payments)").font = bold_font
        ws.cell(row=row, column=2, value=cash_in).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1
        ws.cell(row=row, column=1, value="Cash Out (Expenses)").font = bold_font
        ws.cell(row=row, column=2, value=cash_out).number_format = '#,##0.00'
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        row += 1
        ws.cell(row=row, column=1, value="Net Cash Flow").font = bold_font
        net_cell = ws.cell(row=row, column=2, value=net_cash_flow)
        net_cell.number_format = '#,##0.00'
        net_cell.alignment = Alignment(horizontal='right')
        if net_cash_flow >= 0:
            net_cell.font = Font(bold=True, color="22C55E")
        else:
            net_cell.font = Font(bold=True, color="EF4444")

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="cash_flow_{date_from}_{date_to}.xlsx"'
        )
        wb.save(response)
        return response


class ExportCashFlowPDFView(LoginRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from company_settings.models import Company

        today = timezone.now().date()
        date_from = request.GET.get('date_from', str(today.replace(day=1)))
        date_to = request.GET.get('date_to', str(today))

        cash_in = Payment.objects.filter(
            payment_date__date__gte=date_from,
            payment_date__date__lte=date_to,
            payment_method='cash'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        cash_out = Expense.objects.filter(
            status='paid',
            payment_method='cash',
            created_at__date__gte=date_from,
            created_at__date__lte=date_to
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        net_cash_flow = cash_in - cash_out

        company = Company.objects.first()
        company_name = company.name if company else "J&N WMS"
        currency = company.currency_symbol if company else "MWK"

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="cash_flow_{date_from}_{date_to}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=16, textColor=colors.HexColor('#1E293B'),
                                     spaceAfter=6)
        elements.append(Paragraph(company_name, title_style))
        elements.append(Paragraph("Cash Flow", styles['Heading2']))
        elements.append(Paragraph(f"{date_from} - {date_to}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        data = [['Category', f'Amount ({currency})']]
        data.append(['Cash In (Payments)', f"{cash_in:,.2f}"])
        data.append(['Cash Out (Expenses)', f"{cash_out:,.2f}"])
        data.append(['Net Cash Flow', f"{net_cash_flow:,.2f}"])

        table = Table(data, colWidths=[10*cm, 6*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table)

        doc.build(elements)
        return response


class ExportBalanceSheetExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from company_settings.models import Company
        from company_settings.services import get_setting

        cash_code = get_setting('PAYROLL_CASH_ACCOUNT', '1000')
        bank_code = '1010'
        paye_code = '2005'
        pension_code = '2006'

        def get_balance(account_code):
            balance = JournalLine.objects.filter(
                account__code=account_code
            ).aggregate(b=Sum('debit') - Sum('credit'))['b'] or Decimal('0.00')
            return balance

        cash = get_balance(cash_code)
        bank = get_balance(bank_code)
        paye_liability = get_balance(paye_code)
        pension_liability = get_balance(pension_code)

        total_assets = cash + bank
        total_liabilities = paye_liability + pension_liability
        equity = total_assets - total_liabilities

        company = Company.objects.first()
        currency = company.currency_symbol if company else "MWK"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:B1')
        ws['A1'] = company.name if company else "J&N WMS"
        ws['A1'].font = title_font
        ws.merge_cells('A2:B2')
        ws['A2'] = "Balance Sheet"

        ws.cell(row=4, column=1, value="Assets").font = bold_font
        ws.cell(row=5, column=1, value="Cash")
        ws.cell(row=5, column=2, value=cash).number_format = '#,##0.00'
        ws.cell(row=5, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=6, column=1, value="Bank")
        ws.cell(row=6, column=2, value=bank).number_format = '#,##0.00'
        ws.cell(row=6, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=7, column=1, value="Total Assets").font = bold_font
        ws.cell(row=7, column=2, value=total_assets).number_format = '#,##0.00'
        ws.cell(row=7, column=2).alignment = Alignment(horizontal='right')

        ws.cell(row=9, column=1, value="Liabilities").font = bold_font
        ws.cell(row=10, column=1, value="PAYE Payable")
        ws.cell(row=10, column=2, value=paye_liability).number_format = '#,##0.00'
        ws.cell(row=10, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=11, column=1, value="Pension Payable")
        ws.cell(row=11, column=2, value=pension_liability).number_format = '#,##0.00'
        ws.cell(row=11, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=12, column=1, value="Total Liabilities").font = bold_font
        ws.cell(row=12, column=2, value=total_liabilities).number_format = '#,##0.00'
        ws.cell(row=12, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=13, column=1, value="Equity").font = bold_font
        ws.cell(row=13, column=2, value=equity).number_format = '#,##0.00'
        ws.cell(row=13, column=2).alignment = Alignment(horizontal='right')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="balance_sheet.xlsx"'
        )
        wb.save(response)
        return response


class ExportBalanceSheetPDFView(LoginRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from company_settings.models import Company
        from company_settings.services import get_setting

        cash_code = get_setting('PAYROLL_CASH_ACCOUNT', '1000')
        bank_code = '1010'
        paye_code = '2005'
        pension_code = '2006'

        def get_balance(account_code):
            balance = JournalLine.objects.filter(
                account__code=account_code
            ).aggregate(b=Sum('debit') - Sum('credit'))['b'] or Decimal('0.00')
            return balance

        cash = get_balance(cash_code)
        bank = get_balance(bank_code)
        paye_liability = get_balance(paye_code)
        pension_liability = get_balance(pension_code)

        total_assets = cash + bank
        total_liabilities = paye_liability + pension_liability
        equity = total_assets - total_liabilities

        company = Company.objects.first()
        company_name = company.name if company else "J&N WMS"
        currency = company.currency_symbol if company else "MWK"

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="balance_sheet.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=16, textColor=colors.HexColor('#1E293B'),
                                     spaceAfter=6)
        elements.append(Paragraph(company_name, title_style))
        elements.append(Paragraph("Balance Sheet", styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))

        data1 = [['Assets', f'Amount ({currency})']]
        data1.append(['Cash', f"{cash:,.2f}"])
        data1.append(['Bank', f"{bank:,.2f}"])
        data1.append(['Total Assets', f"{total_assets:,.2f}"])
        table1 = Table(data1, colWidths=[10*cm, 6*cm])
        table1.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table1)
        elements.append(Spacer(1, 0.5*cm))

        data2 = [['Liabilities & Equity', f'Amount ({currency})']]
        data2.append(['PAYE Payable', f"{paye_liability:,.2f}"])
        data2.append(['Pension Payable', f"{pension_liability:,.2f}"])
        data2.append(['Total Liabilities', f"{total_liabilities:,.2f}"])
        data2.append(['Equity', f"{equity:,.2f}"])
        table2 = Table(data2, colWidths=[10*cm, 6*cm])
        table2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        elements.append(table2)

        doc.build(elements)
        return response